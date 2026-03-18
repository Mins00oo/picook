"""
03_export_to_excel.py
정제된 레시피 + 재료 데이터를 백오피스 업로드용 엑셀로 내보내기

사용법:
  1. 먼저 02_refine_with_claude.py로 refined_recipes.json 생성
  2. 먼저 00_extract_ingredients.py로 ingredients_classified.json 생성
  3. pip install openpyxl
  4. python 03_export_to_excel.py

출력:
  - ingredients_for_upload.xlsx (재료 목록 — 백오피스 일괄등록용)
  - recipes_for_upload.xlsx (레시피 3시트 — 백오피스 일괄등록용)
  - recipes_for_review.xlsx (검수용 4시트 — active/wait 색상 구분)
"""

import json
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl 미설치. 설치 후 실행하세요:")
    print("   pip install openpyxl")
    exit(1)

SEEDS_DIR = Path(__file__).parent.parent / "seeds"
REFINED_FILE = SEEDS_DIR / "refined_recipes.json"
INGREDIENTS_FILE = SEEDS_DIR / "ingredients_classified.json"
OUTPUT_INGREDIENTS = SEEDS_DIR / "ingredients_for_upload.xlsx"
OUTPUT_RECIPES = SEEDS_DIR / "recipes_for_upload.xlsx"
OUTPUT_REVIEW = SEEDS_DIR / "recipes_for_review.xlsx"

# 카테고리 → category_id 매핑
CATEGORY_ID_MAP = {
    "채소": 1, "과일": 2, "육류": 3, "해산물": 4,
    "유제품/계란": 5, "곡류/면": 6, "양념/소스": 7, "기타": 8
}

# 재료 아이콘 매핑 (대표적인 것만)
INGREDIENT_ICONS = {
    "양파": "🧅", "마늘": "🧄", "감자": "🥔", "당근": "🥕", "고추": "🌶️",
    "토마토": "🍅", "옥수수": "🌽", "버섯": "🍄", "배추": "🥬", "브로콜리": "🥦",
    "사과": "🍎", "레몬": "🍋", "바나나": "🍌", "포도": "🍇", "딸기": "🍓",
    "수박": "🍉", "복숭아": "🍑", "배": "🍐", "귤": "🍊",
    "소고기": "🥩", "돼지고기": "🥓", "닭고기": "🍗", "오리": "🦆",
    "새우": "🦐", "오징어": "🦑", "생선": "🐟", "조개": "🐚", "게": "🦀",
    "달걀": "🥚", "계란": "🥚", "우유": "🥛", "치즈": "🧀", "버터": "🧈",
    "쌀": "🍚", "밀가루": "🌾", "빵": "🍞", "면": "🍜", "떡": "🍡",
    "소금": "🧂", "꿀": "🍯",
    "두부": "⬜", "물": "💧", "얼음": "🧊",
}

# 스타일
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
ACTIVE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WAIT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def apply_border(ws, max_row: int, max_col: int):
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER


def load_refined() -> list:
    if not REFINED_FILE.exists():
        print(f"❌ {REFINED_FILE} 없음. 02_refine_with_claude.py를 먼저 실행하세요.")
        exit(1)
    with open(REFINED_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def load_ingredients() -> list:
    if not INGREDIENTS_FILE.exists():
        print(f"⚠️ {INGREDIENTS_FILE} 없음. 재료 엑셀은 레시피에서 추출합니다.")
        return []
    with open(INGREDIENTS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def get_icon(name: str) -> str:
    """재료명에 맞는 아이콘 반환."""
    if name in INGREDIENT_ICONS:
        return INGREDIENT_ICONS[name]
    for key, icon in INGREDIENT_ICONS.items():
        if key in name:
            return icon
    return ""


# ── ingredients_for_upload.xlsx ──

def create_ingredients_upload(ingredients: list, recipes: list):
    """백오피스 재료 일괄등록용 엑셀."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "재료 목록"

    headers = ["name", "category_id", "icon", "synonyms"]
    ws.append(headers)
    style_header(ws, len(headers))

    # ingredients_classified.json이 있으면 사용
    if ingredients:
        for ing in ingredients:
            name = ing.get("normalizedName", ing.get("name", ""))
            category = ing.get("category", "기타")
            category_id = CATEGORY_ID_MAP.get(category, 8)
            synonyms = ", ".join(ing.get("synonyms", []))
            icon = get_icon(name)
            ws.append([name, category_id, icon, synonyms])
    else:
        # 없으면 레시피에서 고유 재료 추출
        seen = set()
        for r in recipes:
            for ing in r.get("ingredients", []):
                name = ing.get("name", "")
                if name and name not in seen:
                    seen.add(name)
                    icon = get_icon(name)
                    ws.append([name, 8, icon, ""])  # category_id 기본 8(기타)

    widths = [20, 12, 8, 40]
    for i, w in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

    apply_border(ws, ws.max_row, len(headers))

    SEEDS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_INGREDIENTS)
    count = ws.max_row - 1
    print(f"  🥬 재료 업로드 엑셀: {count}건 → {OUTPUT_INGREDIENTS}")
    return count


# ── recipes_for_upload.xlsx ──

def create_recipes_upload(recipes: list):
    """백오피스 레시피 일괄등록용 엑셀 (3시트)."""
    wb = openpyxl.Workbook()

    # 시트1: 레시피 기본정보
    ws1 = wb.active
    ws1.title = "레시피 기본정보"
    h1 = ["title", "category", "difficulty", "cooking_time_minutes", "servings", "description", "image_url"]
    ws1.append(h1)
    style_header(ws1, len(h1))

    for r in recipes:
        ws1.append([
            r.get("title", ""),
            r.get("category", "korean"),
            r.get("difficulty", "medium"),
            r.get("cookingTimeMinutes", 0),
            r.get("servings", 2),
            r.get("tips", "") or "",
            r.get("mainImage", "") or "",
        ])

    widths1 = [25, 12, 10, 18, 10, 50, 50]
    for i, w in enumerate(widths1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w
    apply_border(ws1, len(recipes) + 1, len(h1))
    print(f"  📋 레시피 기본정보 시트: {len(recipes)}행")

    # 시트2: 레시피-재료 매핑
    ws2 = wb.create_sheet("레시피-재료 매핑")
    h2 = ["recipe_title", "ingredient_name", "amount", "unit", "is_required"]
    ws2.append(h2)
    style_header(ws2, len(h2))

    ing_count = 0
    for r in recipes:
        title = r.get("title", "")
        for ing in r.get("ingredients", []):
            ws2.append([
                title,
                ing.get("name", ""),
                ing.get("amount", 0),
                ing.get("unit", ""),
                "true" if ing.get("isRequired", True) else "false",
            ])
            ing_count += 1

    widths2 = [25, 20, 10, 10, 12]
    for i, w in enumerate(widths2):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w
    apply_border(ws2, ing_count + 1, len(h2))
    print(f"  🥬 레시피-재료 매핑 시트: {ing_count}행")

    # 시트3: 조리 단계
    ws3 = wb.create_sheet("조리 단계")
    h3 = ["recipe_title", "step_number", "description", "step_type", "duration_seconds"]
    ws3.append(h3)
    style_header(ws3, len(h3))

    step_count = 0
    for r in recipes:
        title = r.get("title", "")
        for step in r.get("steps", []):
            ws3.append([
                title,
                step.get("stepNumber", 0),
                step.get("description", ""),
                step.get("stepType", "active"),
                step.get("durationSeconds", 60),
            ])
            step_count += 1

    widths3 = [25, 12, 60, 12, 16]
    for i, w in enumerate(widths3):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w
    apply_border(ws3, step_count + 1, len(h3))
    print(f"  👨‍🍳 조리 단계 시트: {step_count}행")

    SEEDS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_RECIPES)
    print(f"  💾 레시피 업로드 엑셀 → {OUTPUT_RECIPES}")
    return len(recipes), ing_count, step_count


# ── recipes_for_review.xlsx (검수용) ──

def create_review_excel(recipes: list):
    """검수용 엑셀 (active/wait 색상 구분 포함)."""
    wb = openpyxl.Workbook()

    # 시트1: 레시피 요약
    ws1 = wb.active
    ws1.title = "레시피"
    h1 = ["No", "요리명", "카테고리", "난이도", "시간(분)", "인분", "단계수", "재료수", "원본SEQ"]
    ws1.append(h1)
    style_header(ws1, len(h1))

    for i, r in enumerate(recipes):
        ws1.append([
            i + 1, r.get("title", ""), r.get("category", ""), r.get("difficulty", ""),
            r.get("cookingTimeMinutes", ""), r.get("servings", ""),
            len(r.get("steps", [])), len(r.get("ingredients", [])),
            r.get("_sourceSeq", ""),
        ])
    apply_border(ws1, len(recipes) + 1, len(h1))

    # 시트2: 조리단계 (색상 구분)
    ws2 = wb.create_sheet("조리단계")
    h2 = ["레시피No", "요리명", "단계", "설명", "타입", "시간(초)", "병렬가능"]
    ws2.append(h2)
    style_header(ws2, len(h2))

    row_num = 2
    for i, r in enumerate(recipes):
        for step in r.get("steps", []):
            step_type = step.get("stepType", "active")
            ws2.append([
                i + 1, r.get("title", ""), step.get("stepNumber", ""),
                step.get("description", ""), step_type,
                step.get("durationSeconds", ""),
                "Y" if step.get("canParallel", True) else "N",
            ])
            fill = ACTIVE_FILL if step_type == "active" else WAIT_FILL
            for col in range(1, len(h2) + 1):
                cell = ws2.cell(row=row_num, column=col)
                cell.fill = fill
                cell.border = THIN_BORDER
            row_num += 1

    for row in ws2.iter_rows(min_row=2, max_row=row_num - 1, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # 시트3: 요약 통계
    ws3 = wb.create_sheet("요약통계")
    categories = {}
    difficulties = {}
    active_steps = 0
    wait_steps = 0
    for r in recipes:
        cat = r.get("category", "?")
        diff = r.get("difficulty", "?")
        categories[cat] = categories.get(cat, 0) + 1
        difficulties[diff] = difficulties.get(diff, 0) + 1
        for step in r.get("steps", []):
            if step.get("stepType") == "active":
                active_steps += 1
            else:
                wait_steps += 1

    ws3.append(["항목", "값"])
    ws3.append(["총 레시피", len(recipes)])
    ws3.append(["총 조리단계", active_steps + wait_steps])
    ws3.append(["active 단계", active_steps])
    ws3.append(["wait 단계", wait_steps])
    ws3.append([])
    ws3.append(["카테고리", "건수"])
    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        ws3.append([cat, count])
    ws3.append([])
    ws3.append(["난이도", "건수"])
    for diff, count in sorted(difficulties.items(), key=lambda x: -x[1]):
        ws3.append([diff, count])

    wb.save(OUTPUT_REVIEW)
    print(f"  📊 검수용 엑셀 → {OUTPUT_REVIEW}")


if __name__ == "__main__":
    recipes = load_refined()
    ingredients = load_ingredients()
    print(f"📂 정제 레시피 {len(recipes)}건 로드")
    if ingredients:
        print(f"📂 분류 재료 {len(ingredients)}건 로드")
    print()

    # 1. 재료 업로드 엑셀
    print("━━━ ingredients_for_upload.xlsx ━━━")
    ing_total = create_ingredients_upload(ingredients, recipes)

    # 2. 레시피 업로드 엑셀
    print("\n━━━ recipes_for_upload.xlsx ━━━")
    r_count, ri_count, s_count = create_recipes_upload(recipes)

    # 3. 검수용 엑셀
    print("\n━━━ recipes_for_review.xlsx (검수용) ━━━")
    create_review_excel(recipes)

    print(f"\n{'=' * 50}")
    print(f"✅ 완료!")
    print(f"   재료: {ing_total}종")
    print(f"   레시피: {r_count}개 (재료매핑 {ri_count}행, 단계 {s_count}행)")
    print(f"\n📁 생성 파일:")
    print(f"   {OUTPUT_INGREDIENTS}")
    print(f"   {OUTPUT_RECIPES}")
    print(f"   {OUTPUT_REVIEW}")
    print(f"\n📝 다음 단계:")
    print(f"   1. recipes_for_review.xlsx로 검수 (active/wait, duration 확인)")
    print(f"   2. 검수 후 ingredients_for_upload.xlsx → 백오피스 재료 일괄등록")
    print(f"   3. recipes_for_upload.xlsx → 백오피스 레시피 일괄등록")
